import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

directories = {
    r"C:\Reports\Monthly\IMER": 8,
    r"C:\Reports\Monthly\MDS": 9,
    r"C:\Reports\Monthly\PEPFAR_MER_QUARTERLY_MONTHLY": 8,
    r"C:\Reports\Monthly\PEPFAR_MER_SEMI_ANNUAL_MONTHLY": 8,
    r"C:\Reports\Monthly\RMDAH": 9,
    r"C:\Reports\Monthly\RMPREP": 8,
    r"C:\Reports\Monthly\TPT": 7,
    r"C:\Reports\Monthly\TX_TB": 12,
    r"C:\Reports\Quarterly\PEPFAR_MER_QUARTERLY": 8,
    r"C:\Reports\Semi-annual\PEPFAR_MER_SEMI_ANNUAL": 8
}

ignore_sheets = {
    r"C:\Reports\Monthly\PEPFAR_MER_QUARTERLY_MONTHLY": ["PrEP Extra Dissag"],
    r"C:\Reports\Monthly\PEPFAR_MER_SEMI_ANNUAL_MONTHLY": ["TX_TB"],
    r"C:\Reports\Quarterly\PEPFAR_MER_QUARTERLY": ["PrEP Extra Dissag"]
}

def remove_existing_merged_files(directory):
    for filename in os.listdir(directory):
        if "Merged" in filename:
            file_path = os.path.join(directory, filename)
            os.remove(file_path)
            print(f"Removed existing merged file: {file_path}")

def copy_cell_style(source_cell, target_cell):
    target_cell.font = Font(
        name=source_cell.font.name,
        size=source_cell.font.size,
        bold=source_cell.font.bold,
        italic=source_cell.font.italic,
        color=source_cell.font.color
    )
    target_cell.alignment = Alignment(
        horizontal=source_cell.alignment.horizontal,
        vertical=source_cell.alignment.vertical,
        wrap_text=source_cell.alignment.wrap_text
    )
    if source_cell.border:
        target_cell.border = Border(
            left=Side(style=source_cell.border.left.style),
            right=Side(style=source_cell.border.right.style),
            top=Side(style=source_cell.border.top.style),
            bottom=Side(style=source_cell.border.bottom.style)
        )

def remove_blank_rows(ws):
    rows_to_delete = []
    for row in ws.iter_rows():
        if all(cell.value is None for cell in row):
            rows_to_delete.append(row[0].row)
    
    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx)

def remove_first_sheet_if_needed(directory, wb):
    if directory == r"C:\Reports\Monthly\RMDAH":
        first_sheet = wb.sheetnames[0]
        del wb[first_sheet]
        print(f"Removed first sheet: {first_sheet} from {directory}")

def remove_specific_sheet(directory, wb, sheet_name):
    if directory == r"C:\Reports\Monthly\RMDAH" and sheet_name in wb.sheetnames:
        del wb[sheet_name]
        print(f"Removed sheet: {sheet_name} from {directory}")

def merge_files(directory, start_row, ignore_sheets, output_file):
    wb_output = None
    first_file = True

    for filename in os.listdir(directory):
        file_path = os.path.join(directory, filename)
        if file_path.endswith('.xlsx'):
            print(f"Processing file: {file_path}") 
            wb = openpyxl.load_workbook(file_path, data_only=False)
            remove_specific_sheet(directory, wb, "RM de DAH - vertical")
            if first_file:
                wb_output = wb
                first_file = False
            else:
                for sheet_name in wb.sheetnames:
                    if sheet_name in ignore_sheets:
                        print(f"Ignoring sheet: {sheet_name}") 
                        continue
                    ws_src = wb[sheet_name]
                    if sheet_name in wb_output.sheetnames:
                        ws_dst = wb_output[sheet_name]
                    else:
                        ws_dst = wb_output.create_sheet(sheet_name)

                    max_row_dst = ws_dst.max_row
                    for row in ws_src.iter_rows(min_row=start_row, values_only=False):
                        if any(cell.value for cell in row):
                            ws_dst.append([cell.value for cell in row])
                            for src_cell, dst_cell in zip(row, ws_dst[max_row_dst + 1]):
                                copy_cell_style(src_cell, dst_cell)
                            max_row_dst += 1

    if wb_output:
        for sheet in wb_output.worksheets:
            remove_blank_rows(sheet)
        wb_output.save(output_file)
        print(f"Files in {directory} merged into {output_file}")

if __name__ == "__main__":
    for directory, start_row in directories.items():
        if not os.path.exists(directory):
            print(f"Directory {directory} does not exist. Skipping...")
            continue
        remove_existing_merged_files(directory)
        ignore = ignore_sheets.get(directory, [])
        output_file = os.path.join(directory, "Merged_output.xlsx")
        if directory == r"C:\Reports\Monthly\IMER":
            merge_files(directory, 8, ignore, output_file)
        else:
            merge_files(directory, start_row, ignore, output_file)
